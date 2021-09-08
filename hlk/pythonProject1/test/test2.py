# -*- coding: utf-8 -*-
import pandas as pd
from sqlalchemy import create_engine
import time
import xlrd
import openpyxl as op
from openpyxl import load_workbook
import win32com.client as win32


def query():
    # 初始化数据库连接，使用pymysql模块
    # MySQL的用户：root, 密码:147369, 端口：3306,数据库：test
    engine = create_engine('mysql+pymysql://root:1qaz@wsx@192.168.3.202:3306/jxdb_kdxx')
    return engine


def returnstr(str1):  # 格式化文件路径
    str2 = "D:\\\\XZ\\\\"
    str3 = str1.replace("\\", "\\\\")
    str4 = str3.replace("./", "")
    str5 = str2 + str4

    return str5


if __name__ == "__main__":
    start = time.time()

    # engine = query()
    test = []
    wb = load_workbook(r"C:\Users\Hylink\Desktop\summary.xlsx")
    Worksheet = wb["Sheet1"]
    for row in range(1336, Worksheet.max_row):
        print(Worksheet.cell(row=row + 1, column=1).value)#./Excel表格11\0285256201卫裤8.8-8.18 10115条.xls
        str1 = str(Worksheet.cell(row=row + 1, column=1).value)
        print(str1)#./Excel表格11\0285256201卫裤8.8-8.18 10115条.xls
        str2 = returnstr(str1)
        print(str2)#D:\\XZ\\Excel表格11\\0285256201卫裤8.8-8.18 10115条.xls

        wb2 = load_workbook(str2)
        Worksheet2 = wb2["Sheet1"]
        if Worksheet.cell(row=1, column=1) in test:
            print("yes")
        else:
            print("no")
        break
