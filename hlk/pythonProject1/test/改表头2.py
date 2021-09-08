# -*- coding: utf-8 -*-
import time

from openpyxl import load_workbook
import win32com.client as win32
import os


def replace_excel(str5):
    str6 = str5.replace("\\\\", "\\")
    import win32com.client
    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')  # 要看MIME手册
    wb = excel.Workbooks.Open(str6)
    str7 = str6 + "x"
    wb.SaveAs(str7, FileFormat=51)
    wb.Close()
    excel.Application.Quit()
    os.remove(str6)
    return str7


def returnstr(str1):
    str2 = "D:\\\\XZ\\\\"
    # str2 = "C:\\\\Users\\\\Hylink\\\\Desktop\\\\"
    str3 = str1.replace("\\", "\\\\")
    str4 = str3.replace("./", "")
    str5 = str2 + str4
    if str5.find("xlsx") == -1:
        str6 = replace_excel(str5)
    else:
        str6 = str5
    return str6


def replace(str3, setname):
    wb = load_workbook(str3)
    Worksheet = wb[setname]
    # print(Worksheet.cell(row=1, column=1).value)
    for cols1 in range(Worksheet.max_column):
        print(Worksheet.cell(row=1, column=cols1 + 1).value, end=" ")
        if Worksheet.cell(row=1, column=cols1 + 1).value == "工作单号":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递单号"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "物流订单号(唯一标识)":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递单号"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "运单号码":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递单号"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "订单编号":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递单号"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "订单号":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递单号"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "单号":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递单号"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "详情单号":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递单号"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "工作单号":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递单号"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\'运单号":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递单号"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "运单号":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递单号"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "运送单号码":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递单号"

            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "面单号（必填）":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件客户编码"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "'寄方客户编码":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件客户编码"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "寄方客户编码":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件客户编码"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "客户代号":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件客户编码"

            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\nDMS接收时间\n":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件时间"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "时间":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件时间"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "'寄件时间":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件时间"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "派件时间":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件时间"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "下单时间":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件时间"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "创建时间":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件时间"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "邮寄时间":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件时间"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "寄件日期":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件时间"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "提交日期":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件时间"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "日期":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件时间"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "交际日期":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件时间"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\"日期\"":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件时间"
            print("修改成功")

        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\n宅配承运商\n":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "发件人（必填）":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "发件人":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "寄件姓名":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "客服":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "'寄件人":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "寄件人姓名":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "寄件联系人":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人"
            print("修改成功")


        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\n仓库名称\n":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人公司"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "寄件公司":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人公司"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "'寄件公司":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人公司"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "寄件单位":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人公司"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "发件人公司名":
            if Worksheet.cell(row=1, column=cols1 + 2).value == "发件人公司名":
                Worksheet.cell(row=1, column=cols1 + 1).value = ""
                print("修改成功")
            else:
                Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人公司"
                print("修改成功")

            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "寄件电话":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人电话"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "'寄件电话":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人电话"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "寄件联系电话":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人电话"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "发件人电话":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人电话"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "客服电话":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人电话"

            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "寄件地址":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人地址"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\"城市\"":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人地址"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "发件地址（必填）":
            Worksheet.cell(row=1, column=cols1 + 1).value = "寄件人地址"

            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\n收件人\n":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "买家姓名（必填）":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "联系人":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收件姓名":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "'收件人":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收件人姓名":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收货人姓名":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "客户名称":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "客户姓名":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收货人":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收件联系人":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == '"会员名"':
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "姓名":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\"姓名\"":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人"

            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\n收件人详细地址\n":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人地址"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收件地址":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人地址"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收件地址":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人地址"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收件详细地址":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人地址"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收件地点":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人地址"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "买家收货地址（必填）":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人地址"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "详细地址":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人地址"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "地址":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人地址"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\"地区\"":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人地址"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\"地址\"":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人地址"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "联系地址":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人地址"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "客户地址":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人地址"
            print("修改成功")

        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\n商品名称\n":
            Worksheet.cell(row=1, column=cols1 + 1).value = "托寄内容"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "货物名称":
            Worksheet.cell(row=1, column=cols1 + 1).value = "托寄内容"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "'托寄物内容":
            Worksheet.cell(row=1, column=cols1 + 1).value = "托寄内容"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "托寄物内容":
            Worksheet.cell(row=1, column=cols1 + 1).value = "托寄内容"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "商品名称(必填)":
            Worksheet.cell(row=1, column=cols1 + 1).value = "托寄内容"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "商品名称":
            Worksheet.cell(row=1, column=cols1 + 1).value = "托寄内容"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "商品名":
            Worksheet.cell(row=1, column=cols1 + 1).value = "托寄内容"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "产品":
            Worksheet.cell(row=1, column=cols1 + 1).value = "托寄内容"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "订购产品":
            Worksheet.cell(row=1, column=cols1 + 1).value = "托寄内容"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "产品名称":
            Worksheet.cell(row=1, column=cols1 + 1).value = "托寄内容"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "物品名称":
            Worksheet.cell(row=1, column=cols1 + 1).value = "托寄内容"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "物品种类":
            if Worksheet.cell(row=1, column=cols1 + 2).value == "物品详细":
                Worksheet.cell(row=1, column=cols1 + 1).value = ""
                Worksheet.cell(row=1, column=cols1 + 2).value = "托寄内容"
                print("修改成功")
            else:
                Worksheet.cell(row=1, column=cols1 + 1).value = "托寄内容"
                print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "手写品名":
            Worksheet.cell(row=1, column=cols1 + 1).value = "托寄内容"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\"产品名\"":
            Worksheet.cell(row=1, column=cols1 + 1).value = "托寄内容"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\"商品名称\"":
            Worksheet.cell(row=1, column=cols1 + 1).value = "托寄内容"
            print("修改成功")

        elif Worksheet.cell(row=1, column=cols1 + 1).value == "电话号":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人电话"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "电话号码":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人电话"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "'收件电话":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人电话"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收件人手机":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人电话"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收件电话":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人电话"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == " 买家手机号码（必填）":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人电话"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "买家电话":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人电话"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "电话":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人电话"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "联系电话":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人电话"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\"手机号1\"":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人电话"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\"联系电话\"":
            Worksheet.cell(row=1, column=cols1 + 1).value = "收件人电话"

            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\n代收货款\n":
            Worksheet.cell(row=1, column=cols1 + 1).value = "代收金额"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "代收款":
            Worksheet.cell(row=1, column=cols1 + 1).value = "代收金额"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "代收货款":
            Worksheet.cell(row=1, column=cols1 + 1).value = "代收金额"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "代收价格":
            Worksheet.cell(row=1, column=cols1 + 1).value = "代收金额"
            print("修改成功")


        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\n订单数量\n":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递件数"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "件数":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递件数"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == " 数量（必填）":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递件数"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "托寄物数量":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递件数"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "商品数量":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递件数"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "数量":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递件数"

            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "重量":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递重量"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "计费重量(千克)":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递重量"

            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\n商品单价\n":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递金额"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "应收费用":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递金额"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "订单金额":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递金额"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\"金额\"":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递金额"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "金额":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递金额"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "价格":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递金额"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\"价格\"":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递金额"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "实收总资费(元)":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递金额"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "总金额":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递金额"
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "售价":
            Worksheet.cell(row=1, column=cols1 + 1).value = "快递金额"

            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "签收时间":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\n仓库名称\n":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\n收件（省）\n":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\n收件（市）\n":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\n收件（区/县）\n":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\n街道\n":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收件人单位":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收件人所在地":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收件人邮编":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "买家收货省（必填）":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "买家收货市（必填）":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "买家收货区（必填）":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "发件人省（必填）":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "发件人市（必填）":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "发件人区（必填）":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "寄方区号":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "寄件人所在地":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "'寄件手机":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "'收件手机":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "寄件手机":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收件手机":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "收件公司":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "寄件人邮编":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "发件人邮编（必填）":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "大头笔（必填）":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "实际重量(千克)":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "尺寸":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "长(厘米)":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "宽(厘米)":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "高(厘米)":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "备注":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "录单备注":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "订单状态":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "状态":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "签收时间":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        # elif Worksheet.cell(row=1, column=cols1 + 1).value == "商品类别":
        #     Worksheet.cell(row=1, column=cols1 + 1).value = ""
        #   print("修改成功")
        # elif Worksheet.cell(row=1, column=cols1 + 1).value == "商品类别":
        #     Worksheet.cell(row=1, column=cols1 + 1).value = ""
        #   print("修改成功")
        # elif Worksheet.cell(row=1, column=cols1 + 1).value == "商品类别":
        #     Worksheet.cell(row=1, column=cols1 + 1).value = ""
        # print("修改成功")

        elif Worksheet.cell(row=1, column=cols1 + 1).value == "商品单价":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "商品金额":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "优惠金额":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "付款状态":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "付款方式":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "物流公司":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "完成日期":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "付款时间":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "运送方式":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "省份":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "城市":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "地区":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\"时间\"":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "基本总资费(元)":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "应收总资费(元)":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "出库号码":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "汇款预定金额":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "发货时间":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "签件人":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "区号":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "计数单位":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "快件状态":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "邮件号":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\"订购人\"":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\"手机号2\"":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "\"手机号3\"":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        elif Worksheet.cell(row=1, column=cols1 + 1).value == "付款月结号":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")

        elif Worksheet.cell(row=1, column=cols1 + 1).value == "":
            Worksheet.cell(row=1, column=cols1 + 1).value = ""
            print("修改成功")
        else:
            print("未修改")
    wb.save(str3)
    print("cg")

if __name__ == "__main__":

    # index = []
    # setname = 'Sheet1'
    # for str1 in index:
    #     str3 = returnstr(str1)
    #     replace(str3, setname)
    #     print(str1,"成功")



    str1 = "test\\test.xls"
    str1 = r"Excel表格11\0685740_0(2).xls"
    str1 = str1.replace("D:\\XZ\\", "")
    # setname='工作表1'# Sheet1 工作表1
    setname = 'Sheet1'
    # setname = '39'  # Sheet1 工作表1
    # setname = '酒'  # Sheet1 工作表1
    # setname = '蜂胶'  # Sheet1 工作表1
    # setname = '白酒'  # Sheet1 工作表1

    str3 = returnstr(str1)
    replace(str3, setname)
