# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import win32com.client as win32

def returnstr(str1):
    str2 = "D:\\\\XZ\\\\"
    str3 = str1.replace("\\", "\\\\")
    str4 = str3.replace("./", "")
    str5 = str2 + str4
    if str5.find("xlsx") == -1:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(str5)
        str5 = str5 + "x"
        wb.SaveAs(str5, FileFormat=51)  # FileFormat = 51 is for .xlsx extension
        wb.Close()  # FileFormat = 56 is for .xls extension
        excel.Application.Quit()
    return str5

def replace(str3,setname):
    wb = load_workbook(str3)
    Worksheet = wb[setname]
    print(Worksheet.cell(row=1, column=1).value)
    if Worksheet.cell(row=1, column=1).value == "运单号码":
        if Worksheet.cell(row=1, column=12).value == "应收费用":
            Worksheet.cell(row=1, column=1).value = '快递单号'
            Worksheet.cell(row=1, column=2).value = '寄件时间'
            Worksheet.cell(row=1, column=3).value = '收件人'
            Worksheet.cell(row=1, column=4).value = '收件人电话'
            Worksheet.cell(row=1, column=5).value = '收件人地址'
            Worksheet.cell(row=1, column=6).value = '托寄内容'
            Worksheet.cell(row=1, column=7).value = '寄件人'
            Worksheet.cell(row=1, column=8).value = '寄件人电话'
            Worksheet.cell(row=1, column=9).value = '寄件人公司'
            Worksheet.cell(row=1, column=10).value = '寄件人地址'
            Worksheet.cell(row=1, column=11).value = '快递重量'
            Worksheet.cell(row=1, column=12).value = '快递金额'
            print("yes1")
            wb.save(str3)
    elif Worksheet.cell(row=1, column=1).value == "\nDMS接收时间\n":
        for cols1 in range(Worksheet.max_column):
            if Worksheet.cell(row=1, column=1).value == "\nDMS接收时间\n":
                Worksheet.cell(row=1, column=cols1 + 1).value = '寄件时间'
            elif Worksheet.cell(row=1, column=1).value == "\n订单数量\n":
                Worksheet.cell(row=1, column=cols1 + 1).value = '快递件数'
            elif Worksheet.cell(row=1, column=1).value == "\n代收货款\n":
                Worksheet.cell(row=1, column=cols1 + 1).value = '代收金额'
            elif Worksheet.cell(row=1, column=1).value == "\n收件人\n":
                Worksheet.cell(row=1, column=cols1 + 1).value = '收件人'
            elif Worksheet.cell(row=1, column=1).value == "\n收件（省）\n":
                Worksheet.cell(row=1, column=cols1 + 1).value = ''
            elif Worksheet.cell(row=1, column=1).value == "\n收件（市）\n":
                Worksheet.cell(row=1, column=cols1 + 1).value = ''
            elif Worksheet.cell(row=1, column=1).value == "\n收件（区/县）\n":
                Worksheet.cell(row=1, column=cols1 + 1).value = ''
            elif Worksheet.cell(row=1, column=1).value == "\n街道\n":
                Worksheet.cell(row=1, column=cols1 + 1).value = ''
            elif Worksheet.cell(row=1, column=1).value == "\n收件人详细地址\n":
                Worksheet.cell(row=1, column=cols1 + 1).value = '收件人地址'
            elif Worksheet.cell(row=1, column=1).value == "\n商品名称\n":
                Worksheet.cell(row=1, column=cols1 + 1).value = '托寄内容'
            elif Worksheet.cell(row=1, column=1).value == "\n商品单价\n":
                Worksheet.cell(row=1, column=cols1 + 1).value = '快递金额'
            elif Worksheet.cell(row=1, column=1).value == "\n电话号\n":
                Worksheet.cell(row=1, column=cols1 + 1).value = '收件人电话'
            elif Worksheet.cell(row=1, column=1).value == "电话号":
                Worksheet.cell(row=1, column=cols1 + 1).value = '收件人电话'
            elif Worksheet.cell(row=1, column=1).value == "\n\n":
                Worksheet.cell(row=1, column=cols1 + 1).value = ''
            print("yes2")
            wb.save(str3)

    if Worksheet.cell(row=1, column=1).value == "单号":
        for cols1 in range(Worksheet.max_column):
            if Worksheet.cell(row=1, column=1).value == "单号":
                Worksheet.cell(row=1, column=cols1 + 1).value = '快递单号'
            elif Worksheet.cell(row=1, column=1).value == "电话":
                Worksheet.cell(row=1, column=cols1 + 1).value = '收件人电话'
            elif Worksheet.cell(row=1, column=1).value == "地址":
                Worksheet.cell(row=1, column=cols1 + 1).value = '收件人地址'
            elif Worksheet.cell(row=1, column=1).value == "寄件电话":
                Worksheet.cell(row=1, column=cols1 + 1).value = '寄件人电话'
            elif Worksheet.cell(row=1, column=1).value == "寄件公司":
                Worksheet.cell(row=1, column=cols1 + 1).value = '寄件人公司'
            elif Worksheet.cell(row=1, column=1).value == "寄件地址":
                Worksheet.cell(row=1, column=cols1 + 1).value = '寄件人地址'
            elif Worksheet.cell(row=1, column=1).value == "货物名称":
                Worksheet.cell(row=1, column=cols1 + 1).value = '托寄内容'
            elif Worksheet.cell(row=1, column=1).value == "件数":
                Worksheet.cell(row=1, column=cols1 + 1).value = '快递件数'
            elif Worksheet.cell(row=1, column=1).value == "代收款":
                Worksheet.cell(row=1, column=cols1 + 1).value = '代收金额'
            print("yes3")
            wb.save(str3)


# str3 = "C:\\Users\\Hylink\\Desktop\\test\\表头.xlsx"
str1 = r"Excel表格11\6088891-4月卖(1).xls"
setname='Sheet1'

str3=returnstr(str1)
replace(str3,setname)


