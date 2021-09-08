import 改
import 录
import time
from openpyxl import load_workbook
from sqlalchemy import create_engine

if __name__ == "__main__":
    test = ["快递单号", "寄件客户编码", "寄件时间", "寄件人", "寄件人公司", "寄件人电话", "寄件人地址", "收件人", "收件人地址", "托寄内容", "收件人电话", "代收金额", "快递件数",
            "快递重量", "快递金额"]
    start = time.time()
    index = [r"Excel表格1111\6819370_0.xls",
             r"JXJY-2017-038JC-UP-003Excel表格\熙 6月 22.xls"
             # ,r""
             # ,r""
             # ,r""
             ]
    setname = 'Sheet1'
    filename = r'D:\XZ\test\test.xlsx'


    for str1 in index:
        str3 = './' + str1
        str2 = 录.returnstr(str1)
        if str2.find("xlsx") != -1:
            str2 = str2
        elif str1.find("XLSx") != -1:
            str2 = str2
        else:
            str2 = 改.replace_excel(str2)
        wb = load_workbook(str2)  # 加载表格
        sh_name = wb.sheetnames  # 获取所有sheet
        print(sh_name)
        Worksheet2 = wb[sh_name[0]]
        # sh.title = "Sheet1"  # 修改第一个sheet名为dddd
        # wb.save(filename)  # 保存变更
        # wb.close()
        # wb2 = load_workbook(str2)
        # Worksheet2 = wb2[setname]
        print()
        if Worksheet2.cell(row=1, column=1).value in test:
            print("不需要修改")
            # str2 = 录.returnstr(str1)
        else:
            print("需要修改")
            # str2 = 改.returnstr(str1)
            改.replace(str2, setname)
            print(str2, "修改成功")
        engine = 录.query()
        录.insql(str2, engine, setname)
        # print(str2,"录入")
        sql = 'SELECT count(1) FROM `t_kdxx_all_2`;'
        n = 录.queryGp(sql, engine)
        录.write(str3, n)
        print(str1, "成功")
    end = time.time()
    print("执行时间", end - start)
