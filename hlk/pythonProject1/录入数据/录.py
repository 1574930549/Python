# -*- coding: utf-8 -*-
import pandas as pd
from sqlalchemy import create_engine
import xlrd
import openpyxl as op
from openpyxl import load_workbook
import win32com.client as win32

'''
快递单号，寄件客户编码,寄件时间，寄件人，寄件人公司，寄件人电话，寄件人手机，寄件人地址，收件人，收件人地址，托寄内容，收件人电话，代收金额，快递件数，快递重量，快递金额
'''


def query():
    print("querying...")
    # 初始化数据库连接，使用pymysql模块
    # MySQL的用户：root, 密码:147369, 端口：3306,数据库：test
    engine = create_engine('mysql+pymysql://root:1qaz@wsx@192.168.3.202:3306/jxdb_kdxx')
    print("queryover")
    return engine


def returnstr(str1):  # 格式化文件路径
    print("returnstring...")
    str2 = "D:\\\\XZ\\\\"
    str3 = str1.replace("\\", "\\\\")
    str4 = str3.replace("./", "")
    str5 = str2 + str4

    print("returnstrover")
    return str5


def insql(str2, engine, setname):  # 录入数据库
    print("insql")
    dataset = pd.read_excel(str2, header=0, sheet_name=setname)  # Sheet1 工作表1
    columns_names = dataset.columns.values.tolist()
    print(columns_names)
    data = xlrd.open_workbook(str2)  # 文件名以及路径，如果路径或者文件名有中文给前面加一个r拜师原生字符。
    table = data.sheet_by_name(setname)  # 通过名称获取
    index = []
    rows = table.row_values(0)
    print(rows)
    for i in range(len(rows)):
        if rows[i] != '':
            index.append(rows[i])
    dataset = dataset[index]
    print(dataset)
    print("insqling...")
    con = engine.connect()  # 创建连接
    dataset.to_sql(name='t_kdxx_all_2', con=con, if_exists='append', index=False)
    print("insqlover")


def queryGp(sql, engine):  # 查询数据库
    print("queryGping...")
    df = pd.read_sql(sql, con=engine, parse_dates=True)
    str1 = str(df)
    str2 = str1.replace("count(1)", "")
    str3 = str2.replace("0  ", "", 1)
    str4 = str3.replace('\n', '').replace('   ', '')
    print(str4)
    print("queryGpover")
    return str4


def write(name, n):  # 查询后数值写进summary.xlsx
    print("writeing...")
    wbb = xlrd.open_workbook(r"C:\Users\Hylink\Desktop\summary.xlsx")
    Table = wbb.sheet_by_name("Sheet1")
    # Table = workbook.sheet_by_index(0)

    length = Table.nrows
    for i in range(length):
        # print("i")
        row = Table.row_values(i)
        if name in row[0]:
            wb = op.load_workbook(r"C:\Users\Hylink\Desktop\summary.xlsx")
            sh = wb["Sheet1"]
            sh.cell(row=i + 1, column=4, value='已导入')
            sh.cell(row=i + 1, column=5, value=n)

            wb.save(r"C:\Users\Hylink\Desktop\summary.xlsx")
            print("更新成功")
            break
        # else:
        #     print("更新失败，请手动查看")
    print("writeover")