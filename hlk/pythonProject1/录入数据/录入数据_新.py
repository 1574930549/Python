# coding:utf-8
import time
from sqlalchemy import create_engine
import xlrd
import openpyxl as op
import pandas as pd
import 录
def query():
    print("querying...")
    # 初始化数据库连接，使用pymysql模块
    # MySQL的用户：root, 密码:147369, 端口：3306,数据库：test
    engine = create_engine('mysql+pymysql://root:1qaz@wsx@192.168.3.202:3306/jxdb_kdxx')
    print("queryover")
    return engine

def returnstr(str1):  # 格式化文件路径
    print("returnstring...")
    str2 = "D:\\\\XZ\\\\"
    str3 = str1.replace("\\", "\\\\")
    str4 = str3.replace("./", "")
    str5 = str2 + str4

    print("returnstrover")
    return str5
def change2(data):
    data.rename(
        columns={"工作单号": "快递单号", "物流订单号(唯一标识)": "快递单号", "运单号码": "快递单号", "订单编号": "快递单号", "订单号": "快递单号",
                 "\n订单号\n": "快递单号",
                 "单号": "快递单号", "详情单号": "快递单号", " 工作单号": "快递单号", "\'运单号": "快递单号", "运单号": "快递单号", "运送单号码": "快递单号",
                 "面单号（必填）": "寄件客户编码", "'寄方客户编码": "寄件客户编码", "寄方客户编码": "寄件客户编码", "客户代号": "寄件客户编码", "\nDMS接收时间\n": "寄件时间",
                 "DMS接收时间": "寄件时间", "时间": "寄件时间", "'寄件时间": "寄件时间", "派件时间": "寄件时间", "下单时间": "寄件时间", "创建时间": "寄件时间",
                 "实际出货时间": "寄件时间", "邮寄时间": "寄件时间", "订单时间": "寄件时间", "订单时间 ": "寄件时间", "寄件日期": "寄件时间", "提交日期": "寄件时间",
                 "日期": "寄件时间", "交际日期": "寄件时间", "\"日期\"": "寄件时间", "\n宅配承运商\n": "寄件人", "发件人（必填）": "寄件人", "发件人": "寄件人",
                 "寄件姓名": "寄件人", "客服": "寄件人", "'寄件人": "寄件人", "寄件人姓名": "寄件人", "寄件联系人": "寄件人", "宅配承运商": "寄件人",
                 "发货地点": "寄件人公司", "寄件公司": "寄件人公司", "'寄件公司": "寄件人公司", "寄件单位": "寄件人公司",
                 "寄件公司名称": "寄件人公司", "发件人公司名": "寄件人公司", "寄件电话": "寄件人电话", "'寄件电话": "寄件人电话", "寄件联系电话": "寄件人电话",
                 "寄件联系方式": "寄件人电话", "发件人电话": "寄件人电话", "客服电话": "寄件人电话", "寄件地址": "寄件人地址", "\"城市\"": "寄件人地址",
                 "发件地址（必填）": "寄件人地址", "寄件详细地址": "寄件人地址", "\n收件人\n": "收件人", "买家姓名（必填）": "收件人", "联系人": "收件人",
                 "收件姓名": "收件人",
                 "'收件人": "收件人", "收件人姓名": "收件人", "收货人姓名": "收件人", "客户名称": "收件人", "客户姓名": "收件人", "收货人": "收件人",
                 "收件联系人": "收件人",
                 '"会员名"': "收件人", "姓名": "收件人", "\"姓名\"": "收件人", "\n收件人详细地址\n": "收件人地址", "收件人详细地址": "收件人地址",
                 "收件地址": "收件人地址",
                 " 收件地址": "收件人地址", "收件详细地址": "收件人地址", "收件地点": "收件人地址", "买家收货地址（必填）": "收件人地址", "详细地址": "收件人地址",
                 "地址": "收件人地址",
                 "\"地区\"": "收件人地址", "\"地址\"": "收件人地址", "联系地址": "收件人地址", "客户地址": "收件人地址", "\n商品名称\n": "托寄内容",
                 "货物名称": "托寄内容",
                 "'托寄物内容": "托寄内容", "托寄物内容": "托寄内容", "商品名称(必填)": "托寄内容", "商品名称": "托寄内容", "商品名": "托寄内容", "产品": "托寄内容",
                 "订购产品": "托寄内容", "产品名称": "托寄内容", "物品名称": "托寄内容", "内件名称": "托寄内容", "物品种类": "", "物品详细": "托寄内容",
                 "手写品名": "托寄内容",
                 "\"产品名\"": "托寄内容", "\"商品名称\"": "托寄内容", "电话号": "收件人电话", "电话号码": "收件人电话", "'收件电话": "收件人电话",
                 "手机": "收件人电话",
                 "收件人手机": "收件人电话", "收件电话": "收件人电话", " 买家手机号码（必填）": "收件人电话", "买家电话": "收件人电话", "电话": "收件人电话",
                 "联系电话": "收件人电话",
                 "\"手机号1\"": "收件人电话", "\"联系电话\"": "收件人电话", "收件联系方式": "收件人电话", "\n代收货款\n": "代收金额", " 代收货款": "代收金额",
                 "代收款": "代收金额", "代收货款": "代收金额", "代收价格": "代收金额", "\n订单数量\n": "快递件数", "订单数量": "快递件数", "件数": "快递件数",
                 " 数量（必填）": "快递件数", "托寄物数量": "快递件数", "商品数量": "快递件数", "数量": "快递件数", "重量": "快递重量", "计费重量(千克)": "快递重量",
                 "\n商品单价\n": "快递金额", "应收费用": "快递金额", "订单金额": "快递金额", "货款金额": "快递金额", "\"金额\"": "快递金额", "金额": "快递金额",
                 "价格": "快递金额", "\"价格\"": "快递金额", "实收总资费(元)": "快递金额", "总金额": "快递金额", "售价": "快递金额", "产品单价": "快递金额",
                 "签收时间": "", "物流快递": "", "\n仓库名称\n": "", "仓库名称": "", "\n收件（省）\n": "", "收件（省）": "", "\n收件（市）\n": "",
                 "收件（市）": "", "\n收件（区/县）\n": "", "收件（区/县）": "", "\n街道\n": "", "街道": "", "收件人单位": "", "收件人所在地": "",
                 "收件人邮编": "", "买家收货省（必填）": "", "买家收货市（必填）": "", "买家收货区（必填）": "", "发件人省（必填）": "", "发件人市（必填）": "",
                 "发件人区（必填）": "", "寄方区号": "", "寄件人所在地": "", "寄件所在地": "", "'寄件手机": "", "'收件手机": "", "寄件手机": "",
                 "收件手机": "",
                 "收件所在地": "", "收件公司": "", "收件公司名称": "", "寄件人邮编": "", "发件人邮编（必填）": "", "\n运单号\n": "", "大头笔（必填）": "",
                 "实际重量(千克)": "", "尺寸": "", "长(厘米)": "", "宽(厘米)": "", "高(厘米)": "", "备注": "", "录单备注": "", "订单状态": "",
                 "状态": "", " 签收时间": "", "商品单价": "", "商品金额": "", "优惠金额": "", "付款状态": "", "付款方式": "", "物流公司": "",
                 " 仓库名称": "",
                 "完成日期": "", "付款时间": "", "运送方式": "", "省份": "", "城市": "", "地区": "", "\"时间\"": "", "基本总资费(元)": "",
                 "应收总资费(元)": "", "出库号码": "", "汇款预定金额": "", "发货时间": "", "签件人": "", "区号": "", "计数单位": "", "快件状态": "",
                 "邮件号": "", "寄件邮编": "", "\"邮编\"": "", "收件邮编": "", "\"订购人\"": "", "\"手机号2\"": "", "\"手机号3\"": "",
                 "付款月结号": ""}, inplace=True)
    return data

def change(str2):
    with pd.ExcelWriter(str2) as writer:
        data_xls = pd.io.excel.ExcelFile(str2)
        data={}
        print(data_xls.sheet_names)
        for name in data_xls.sheet_names:
            df = pd.read_excel(data_xls, sheet_name=name, header=0)
            df = df.loc[:, ~df.columns.str.contains('Unnamed')]
            data[name] = df
            df=change2(df)
            # print(data)
            # print(data[name])
            df.to_excel(writer, index=False, sheet_name=name)

def insql(str2, engine):  # 录入数据库
    print("insql")
    data_xls = pd.io.excel.ExcelFile(str2)
    data = {}
    # print(data_xls.sheet_names)
    for name in data_xls.sheet_names:
        df = pd.read_excel(data_xls, sheet_name=name, header=0)
        columns_names = df.columns.values.tolist()
        # print(columns_names)
        index = []
        for i in columns_names:
            if i != '':
                if i.find("Unnamed:")==-1:
                    index.append(i)
                    print(i)
        dataset = df[index]
        print(dataset)
        print("insqling...")
        con = engine.connect()  # 创建连接
        dataset.to_sql(name='t_kdxx_all_2', con=con, if_exists='append', index=False)
        print("insqlover")


def queryGp(sql, engine):  # 查询数据库
    print("queryGping...")
    df = pd.read_sql(sql, con=engine, parse_dates=True)
    str1 = str(df)
    str2 = str1.replace("count(1)", "")
    str3 = str2.replace("0  ", "", 1)
    str4 = str3.replace('\n', '').replace('   ', '')
    print(str4)
    print("queryGpover")
    return str4

def write(name, n):  # 查询后数值写进summary.xlsx
    print("writeing...")
    wbb = xlrd.open_workbook(r"C:\Users\Hylink\Desktop\summary.xlsx")
    Table = wbb.sheet_by_name("Sheet1")
    # Table = workbook.sheet_by_index(0)

    length = Table.nrows
    for i in range(length):
        # print("i")
        row = Table.row_values(i)
        if name in row[0]:
            wb = op.load_workbook(r"C:\Users\Hylink\Desktop\summary.xlsx")
            sh = wb["Sheet1"]
            sh.cell(row=i + 1, column=4, value='已导入')
            sh.cell(row=i + 1, column=5, value=n)

            wb.save(r"C:\Users\Hylink\Desktop\summary.xlsx")
            print("更新成功")
            break
        # else:
        #     print("更新失败，请手动查看")
    print("writeover")
if __name__ == "__main__":
    start = time.time()
    test=[r"JXJY-2017-038JC-HD-001Excel表格\747e26538837f3c8a5c735d0e62a9644_卫裤最新.xls5000(1).xls",
          r"导出Excel\5-28 2940576.xls",
          r"JXJY-2017-038JC-UP-001Excel表格\中国必胜手表(1).xls"
          ]
    for str1 in test:
        str1=str1.replace("D:\\XZ\\","")
        str3 = './' + str1
        str2 = returnstr(str1)
        change(str2)
        engine = query()
        insql(str2, engine)

        sql = 'SELECT count(1) FROM `t_kdxx_all_2`;'
        n= queryGp(sql, engine)

        write(str3, n)
        print(str3)
    end = time.time()
    print("执行时间", end - start)
