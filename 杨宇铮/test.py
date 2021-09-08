# coding:utf-8
import time
from sqlalchemy import create_engine
import xlrd
import openpyxl as op
import pandas as pd


def DataProcessing(str1):
    data_xlsx = pd.io.excel.ExcelFile(str1)  # 打开这个Excel并把值赋给data_xlsx
    df = pd.read_excel(data_xlsx, sheet_name="Sheet1", header=0)  # 读取文件中Sheet1表中内容（header0代表头为第一行）并把值赋给df
    df_li = df.values.tolist()  # 读取df每一列内容并把值赋给df_li
    Latitude = []  # 定义空列表（维度）
    Longitude = []  # 定义空列表（经度）
    DetectionDate = []
    SubmissionDate = []
    for s_li in df_li:  # 遍历每一列
        DetectionDate.append(s_li[1])
        SubmissionDate.append(s_li[5])
        Latitude.append(s_li[6])  # 维度在第7列
        Longitude.append(s_li[7])  # 经度在第8列
    # print(DetectionDate[0])
    return DetectionDate, SubmissionDate, Latitude, Longitude


def write(DetectionDate):
    dict1 = {DetectionDate[0]: 1}
    for i in range(1, len(DetectionDate)):
        if DetectionDate[i] in dict1:
            j = dict1[DetectionDate[i]]
            dict1[DetectionDate[i]] = j + 1
        else:
            dict1[DetectionDate[i]] = 1
    wb = op.load_workbook(r"G:\代码\Python\杨宇铮\name.xlsx")  # 用openpyxl打开文件
    sh = wb["Sheet1"]  # 打开Sheet表
    a = list(dict1.keys())
    b = list(dict1.values())
    for i in range(len(dict1)):
        sh.cell(row=i + 1, column=1, value=a[i])  # 更改特定行列内容
        sh.cell(row=i + 1, column=2, value=str(b[i]) + '次')  # 更改特定行列内容
    wb.save(r"G:\代码\Python\杨宇铮\name.xlsx")


def DetectionDateup(DetectionDate, Latitude, Longitude,str2):
    wb = op.Workbook()
    wb.save("G:\\代码\\Python\\杨宇铮\\坐标\\"+str2+".xlsx")
    wb = op.load_workbook("G:\\代码\\Python\\杨宇铮\\坐标\\"+str2+".xlsx")  # 用openpyxl打开文件
    sh = wb["Sheet"]  # 打开Sheet表
    j=0
    for i in range(len(DetectionDate)):
        # print(i)
        if str(DetectionDate[i]).find(str2)!=-1:
            # print(DetectionDate[i],end=",")
            # print(Latitude[i],end=",")
            # print(Longitude[i])
            sh.cell(row=j + 1, column=1, value=DetectionDate[i])  # 更改特定行列内容
            sh.cell(row=j + 1, column=2, value=Latitude[i])  # 更改特定行列内容
            sh.cell(row=j + 1, column=3, value=Longitude[i])  # 更改特定行列内容
            j=j+1
    wb.save("G:\\代码\\Python\\杨宇铮\\坐标\\"+str2+".xlsx")
    print(str2,"写入成功")


if __name__ == "__main__":
    str1 = r"G:\代码\Python\杨宇铮\2021MCMProblemC_DataSet(1).xlsx"  # 文件路径
    DetectionDate, SubmissionDate, Latitude, Longitude = DataProcessing(str1)
    # write(DetectionDate)
    # str2="2007"
    # DetectionDateup(DetectionDate, Latitude, Longitude,str2)
    list=['2007','2010','2012','2013','2014','2015','2016','2017','2018','2019-01','2019-03','2019-04','2019-05','2019-06'
        , '2019-07','2019-08','2019-09','2019-10','2019-11','2019-12','2020-01','2020-02','2020-03','2020-04','2020-05'
        , '2020-06','2020-07','2020-08','2020-09','2020-10','2020-11','2020-12']
    for str2 in list:
        DetectionDateup(DetectionDate, Latitude, Longitude, str2)
